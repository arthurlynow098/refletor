import os
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import json

# ------------------ CONFIGURAÇÃO ------------------
# Intervalo de leitura. Se um arquivo falhar, ajuste estes valores.
START_ROW = 2
END_ROW = 14
START_COL = "A"
END_COL = "AF"

# Mapeamento de códigos de texto para nomes de cores (usado no JSON)
COLOR_MAP = {
    'VM': 'red',
    'AZ': 'blue',
    'BR': 'white',
    'AB': 'yellow'
}
# Mapeamento para nomes em português (usado no relatório .txt)
COLOR_NAME_PT = {
    'red': 'vermelho',
    'blue': 'azul',
    'white': 'branco',
    'yellow': 'âmbar'
}
VALID_MODULE_NUMBERS = range(1, 25)
# --------------------------------------------------

def selecionar_arquivo_excel():
    """Abre uma janela para o usuário selecionar um arquivo Excel."""
    root = tk.Tk()
    root.withdraw()
    filetypes = (("Arquivos Excel", "*.xlsx *.xlsm"), ("Todos os arquivos", "*.*"))
    caminho = filedialog.askopenfilename(title="Selecione o arquivo Excel com o layout", filetypes=filetypes)
    return caminho

def escolher_planilha_pol(caminho_arquivo):
    """Lista as planilhas 'POL' e permite que o usuário escolha uma, se houver várias."""
    try:
        wb = load_workbook(caminho_arquivo, read_only=True)
        sheet_names = wb.sheetnames
        pol_sheets = [s for s in sheet_names if "POL" in s.upper()]

        if not pol_sheets:
            messagebox.showinfo("Aviso", f"Nenhuma planilha com 'POL' no nome foi encontrada. Usando a primeira planilha: '{sheet_names[0]}'")
            return sheet_names[0] if sheet_names else None
        
        if len(pol_sheets) == 1:
            return pol_sheets[0]

        # Diálogo para o usuário escolher a planilha
        root = tk.Tk()
        root.withdraw()
        escolha = tk.StringVar(root)
        escolha.set(pol_sheets[0])
        
        popup = tk.Toplevel()
        popup.title("Selecione a Planilha")
        tk.Label(popup, text="Múltiplas planilhas 'POL' encontradas. Qual você deseja usar?").pack(padx=20, pady=10)
        
        option_menu = tk.OptionMenu(popup, escolha, *pol_sheets)
        option_menu.pack(padx=20, pady=5)
        
        planilha_selecionada = None
        def on_ok():
            nonlocal planilha_selecionada
            planilha_selecionada = escolha.get()
            popup.destroy()

        tk.Button(popup, text="Confirmar", command=on_ok).pack(padx=20, pady=10)
        popup.wait_window()
        return planilha_selecionada

    except Exception as e:
        messagebox.showerror("Erro de Leitura", f"Não foi possível ler as planilhas do arquivo:\n{e}")
        return None

def extrair_dados_da_planilha(caminho_arquivo, nome_planilha):
    """Extrai dados do intervalo, preenchendo corretamente as células mescladas."""
    try:
        wb = load_workbook(caminho_arquivo, data_only=True)
        ws = wb[nome_planilha]

        col_start_idx = column_index_from_string(START_COL)
        col_end_idx = column_index_from_string(END_COL)
        
        # Cria uma grade de dados vazia
        grid_data = []
        for r in range(START_ROW, END_ROW + 1):
            row_data = [ws.cell(row=r, column=c).value for c in range(col_start_idx, col_end_idx + 1)]
            grid_data.append(row_data)

        # Preenche os valores das células mescladas
        for merged_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value
            
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    # Verifica se a célula está dentro do nosso intervalo de interesse
                    if START_ROW <= r <= END_ROW and col_start_idx <= c <= col_end_idx:
                        grid_r_idx = r - START_ROW
                        grid_c_idx = c - col_start_idx
                        grid_data[grid_r_idx][grid_c_idx] = top_left_cell_value
        
        return pd.DataFrame(grid_data)
    except Exception as e:
        print(f"Erro ao ler o intervalo do Excel: {e}")
        return None

def encontrar_elementos(df):
    """Encontra todas as ocorrências de números de módulo e células de cor."""
    numeros = []
    cores = []
    
    for r_idx, row in df.iterrows():
        for c_idx, cell_value in enumerate(row):
            if pd.isna(cell_value):
                continue
            
            cell_text = str(cell_value).strip()
            
            # Tenta converter para número
            try:
                num = int(float(cell_text))
                if num in VALID_MODULE_NUMBERS:
                    numeros.append({'value': num, 'r': r_idx, 'c': c_idx})
                    continue 
            except (ValueError, TypeError):
                pass

            # Encontra códigos de cor
            cell_upper = cell_text.upper()
            for codigo, nome_cor in COLOR_MAP.items():
                if codigo in cell_upper:
                    cores.append({'color': nome_cor, 'r': r_idx, 'c': c_idx, 'content': str(cell_value).replace('\n', ' ')})
                    break 
    
    return numeros, cores

def associar_por_proximidade(numeros, cores):
    """Para cada ocorrência de número, encontra sua cor mais próxima."""
    layout_final = []

    for num_data in numeros:
        melhor_cor = None
        distancia_minima = float('inf')
        
        for cor_data in cores:
            distancia = abs(num_data['r'] - cor_data['r']) + abs(num_data['c'] - cor_data['c'])
            
            if distancia < distancia_minima:
                distancia_minima = distancia
                melhor_cor = cor_data['color']
        
        if melhor_cor:
            layout_final.append({
                'module': num_data['value'],
                'color': melhor_cor
            })
            
    return layout_final

def salvar_arquivos_resultado(layout, caminho_arquivo):
    """Salva os resultados em arquivos .txt (legível) e .json (para o site)."""
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    diretorio = os.path.dirname(caminho_arquivo)
    
    caminho_saida_txt = os.path.join(diretorio, f"{nome_base}_layout_extraido.txt")
    caminho_saida_json = os.path.join(diretorio, f"{nome_base}_layout.json")

    layout_ordenado = sorted(layout, key=lambda k: k['module'])
    
    # Salvar em .txt
    with open(caminho_saida_txt, 'w', encoding='utf-8') as f:
        f.write(f"--- Layout de Cores Extraído ---\n")
        f.write(f"Arquivo: {os.path.basename(caminho_arquivo)}\n")
        f.write("="*40 + "\n\n")
        if not layout_ordenado:
            f.write("Nenhum módulo foi associado a uma cor.\n")
        else:
            for item in layout_ordenado:
                cor_pt = COLOR_NAME_PT.get(item['color'], item['color'])
                f.write(f"Módulo {item['module']}: {cor_pt}\n")

    # Salvar em .json
    with open(caminho_saida_json, 'w', encoding='utf-8') as f:
        json.dump(layout_ordenado, f, indent=4)

    print(f"\nLayout legível salvo em: '{caminho_saida_txt}'")
    print(f"Layout JSON para o site salvo em: '{caminho_saida_json}'")

def salvar_debug_data(df, caminho_arquivo):
    """NOVO: Salva os dados brutos lidos em um CSV para depuração."""
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    diretorio = os.path.dirname(caminho_arquivo)
    caminho_saida_csv = os.path.join(diretorio, f"{nome_base}_DADOS_LIDOS_PARA_DEBUG.csv")
    df.to_csv(caminho_saida_csv, index=False, header=False, sep=';')
    print(f"\nArquivo de depuração foi salvo em: '{caminho_saida_csv}'")
    print("Abra este arquivo para ver exatamente o que o script está lendo.")

# --- Execução Principal ---
if __name__ == "__main__":
    arquivo_excel = selecionar_arquivo_excel()

    if not arquivo_excel:
        print("Nenhum arquivo selecionado. Encerrando.")
    else:
        print(f"\nProcessando arquivo: {os.path.basename(arquivo_excel)}")
        planilha_alvo = escolher_planilha_pol(arquivo_excel)

        if planilha_alvo:
            df_dados = extrair_dados_da_planilha(arquivo_excel, planilha_alvo)
            
            if df_dados is not None:
                salvar_debug_data(df_dados, arquivo_excel) # Gera o arquivo de depuração

                numeros_encontrados, cores_encontradas = encontrar_elementos(df_dados)
                print(f"\nAnálise: {len(numeros_encontrados)} módulos (incluindo repetidos) e {len(cores_encontradas)} células de cor encontrados.")
                
                if not numeros_encontrados or not cores_encontradas:
                    messagebox.showerror("Erro de Análise", "Não foi possível encontrar módulos ou cores no intervalo especificado. Verifique o arquivo de depuração e as configurações de intervalo no script.")
                else:
                    print("\nAssociando módulos às cores...")
                    layout = associar_por_proximidade(numeros_encontrados, cores_encontradas)
                    salvar_arquivos_resultado(layout, arquivo_excel)
                    
                    print("\n===========================================")
                    print(" LAYOUT EXTRAÍDO COM SUCESSO!")
                    print("===========================================")
            else:
                messagebox.showerror("Erro de Extração", "Não foi possível extrair dados da planilha.")
                
    input("\nPressione Enter para fechar...")