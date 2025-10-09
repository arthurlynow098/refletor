import openpyxl
import os
import warnings
import re
import json
import sys

# Suprime o aviso específico sobre cabeçalhos/rodapés que não podem ser analisados
warnings.filterwarnings('ignore', category=UserWarning)

def parse_info_string(info_string, patterns):
    """
    Analisa a string de descrição do produto e extrai as informações
    em um dicionário estruturado, tentando múltiplos padrões de regex.
    """
    match = None
    for pattern in patterns:
        match = pattern.search(info_string)
        if match:
            break

    if not match:
        return None

    data = match.groupdict()

    # --- Limpeza e formatação dos dados ---

    # Normaliza os campos que podem ou não existir dependendo do padrão
    data.setdefault('produto', 'N/A')
    data.setdefault('modelo', 'N/A')
    data.setdefault('tamanho', '0P')
    data.setdefault('modulos', '0M')
    data.setdefault('leds', '0L')
    data.setdefault('cores_leds', '()')
    data.setdefault('funcoes_extras', 'N/A')
    data.setdefault('cores_tampas', 'N/A')

    # Processa os campos para o formato final
    data['tamanho_polegadas'] = data.get('tamanho', '0P').replace('P', '')
    data['cores_tampas'] = data.get('cores_tampas', '').strip()
    data['quantidade_modulos'] = data.get('modulos', '0M').replace('M', '')
    data['numero_leds'] = data.get('leds', '0L').replace('L', '')

    # Extrai a lista de cores de forma mais robusta
    cores_str = data.get('cores_leds', '()').strip('()')
    data['cores_leds_lista'] = [c.strip() for c in cores_str.split('+') if c.strip()]

    # Remove as chaves originais que foram substituídas para limpar o JSON final
    # Usamos .pop() com um valor padrão para evitar erros se a chave não existir
    data.pop('tamanho', None)
    data.pop('modulos', None)
    data.pop('leds', None)

    return data

def extrair_numero_da_celula(sheet, coord):
    """
    Extrai o valor numérico de uma única célula.
    """
    if not coord:
        return None
    
    try:
        cell = sheet[coord]
        if cell and cell.value is not None:
            return int(cell.value)
    except (ValueError, TypeError):
        # O valor na célula não é um número
        return None
    except Exception as e:
        print(f"AVISO: Erro ao ler a célula de canal '{coord}': {e}")

    return None


def mapear_cores_posicoes_fixas(sheet):
    """
    Verifica um conjunto de posições fixas na planilha para mapear as cores dos módulos.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): A planilha do Excel a ser analisada.

    Returns:
        list: Uma lista de dicionários, cada um representando um módulo com sua posição e cor.
    """
    COLOR_MAP = {'VM': 'vermelho', 'AZ': 'azul', 'BR': 'branco', 'AB': 'ambar'}

    # Mapeamento do deslocamento (offset) para encontrar o canal relativo a cada módulo.
    # Formato: (row_offset, column_offset)
    POSICOES_CANAIS = {
        # Módulos 1 a 8 (Abaixo)
        "O4:P5": (2, 0), "Q4:R5": (2, 0), "S4:T5": (2, 0), "V4:W5": (2, 0),
        "X4:Y5": (2, 0), "Z4:AA5": (2, 0), "AB6:AC7": (2, 0), "AD10:AE11": (2, 0),
        
        # Módulos 9 a 23 (Acima, Esquerda, Direita)
        "AD14:AE15": (0, -2), # Módulo 9 (Esquerda)
        "D14:E15": (0, 2),    # Módulo 23 (Direita)

        "AD18:AE19": (-2, 0),# Módulo 10 (Acima)
        "AB22:AC23": (-2, 0),# Módulo 11 (Acima)
        "Z24:AA25": (-2, 0), # Módulo 12 (Acima)
        "X24:Y25": (-2, 0),  # Módulo 13 (Acima)
        "V24:W25": (-2, 0),  # Módulo 14 (Acima)
        "S24:T25": (-2, 0),  # Módulo 15 (Acima)
        "Q24:R25": (-2, 0),  # Módulo 16 (Acima)
        "O24:P25": (-2, 0),  # Módulo 17 (Acima)
        "L24:M25": (-2, 0),  # Módulo 18 (Acima)
        "J24:K25": (-2, 0),  # Módulo 19 (Acima)
        "H24:I25": (-2, 0),  # Módulo 20 (Acima)
        "F22:G23": (-2, 0),  # Módulo 21 (Acima)
        "D18:E19": (-2, 0),  # Módulo 22 (Acima)

        # Módulos 24 a 28 (Abaixo)
        "D10:E11": (2, 0), "F6:G7": (2, 0), "H4:I5": (2, 0),
        "J4:K5": (2, 0), "L4:M5": (2, 0),
    }

    posicoes_modulos = [
        "O4:P5", "Q4:R5", "S4:T5", "V4:W5", "X4:Y5", "Z4:AA5", "AB6:AC7", 
        "AD10:AE11", "AD14:AE15", "AD18:AE19", "AB22:AC23", "Z24:AA25", 
        "X24:Y25", "V24:W25", "S24:T25", "Q24:R25", "O24:P25", "L24:M25", 
        "J24:K25", "H24:I25", "F22:G23", "D18:E19", "D14:E15", "D10:E11", 
        "F6:G7", "H4:I5", "J4:K5", "L4:M5"
    ]

    mapeamento_final = []
    for i, range_str in enumerate(posicoes_modulos, 1):
        cores_e_canais = []
        
        try:
            canal_info = POSICOES_CANAIS.get(range_str)

            for row_of_cells in sheet[range_str]:
                for cell in row_of_cells:
                    if not hasattr(cell, 'coordinate'):
                        continue

                    # Encontra cores na célula atual
                    if cell.value and isinstance(cell.value, str):
                        texto_celula = cell.value.strip().upper()
                        cores_na_celula = [
                            nome_cor for codigo, nome_cor in COLOR_MAP.items() 
                            if codigo in texto_celula
                        ]

                        if cores_na_celula:
                            coord_canal_final = None
                            if isinstance(canal_info, tuple): # Lógica de offset relativo
                                row_offset, col_offset = canal_info
                                coord_cor = openpyxl.utils.cell.coordinate_from_string(cell.coordinate) # type: ignore
                                col_cor_idx = openpyxl.utils.column_index_from_string(coord_cor[0])
                                coord_canal_final = openpyxl.utils.get_column_letter(col_cor_idx + col_offset) + str(coord_cor[1] + row_offset)

                            numero_canal = extrair_numero_da_celula(sheet, coord_canal_final)

                            for cor in cores_na_celula:
                                cores_e_canais.append({
                                    "cor": cor, 
                                    "canal": numero_canal,
                                    "canal_encontrado_em": coord_canal_final # Adiciona a célula de depuração
                                })

        except Exception as e:
            print(f"AVISO: Não foi possível processar o intervalo '{range_str}'. Erro: {e}")
        
        # Debug: mostra o que foi encontrado para cada módulo
        if cores_e_canais:
            print(f"DEBUG: Módulo {i} ({range_str}) - Encontradas {len(cores_e_canais)} cores")
        
        mapeamento_final.append({
            "modulo": i,
            "posicao": range_str,
            "mapeamento": cores_e_canais
        })
        
    return mapeamento_final

def contar_modulos_ativos(mapeamento_cores_modulos):
    """
    Conta quantos módulos realmente têm dados (não estão vazios)
    
    Args:
        mapeamento_cores_modulos (list): Lista de módulos com seus mapeamentos
        
    Returns:
        int: Número de módulos que possuem pelo menos uma cor/canal definido
    """
    modulos_ativos = 0
    modulos_com_dados = []
    
    for modulo_info in mapeamento_cores_modulos:
        modulo_num = modulo_info.get('modulo', 0)
        
        # Verifica se o módulo tem pelo menos um mapeamento de cor válido
        if modulo_info.get('mapeamento') and len(modulo_info['mapeamento']) > 0:
            # Verifica se existe pelo menos um mapeamento com cor e canal válidos
            tem_dados_validos = False
            cores_encontradas = []
            
            for mapping in modulo_info['mapeamento']:
                cor = mapping.get('cor')
                canal = mapping.get('canal')
                
                if (cor and cor.strip() and 
                    canal is not None and 
                    canal != 0):
                    tem_dados_validos = True
                    cores_encontradas.append(f"{cor}(canal {canal})")
            
            if tem_dados_validos:
                modulos_ativos += 1
                modulos_com_dados.append({
                    'modulo': modulo_num,
                    'posicao': modulo_info.get('posicao', ''),
                    'cores': cores_encontradas
                })
    
    # Debug: mostra quais módulos foram encontrados
    print(f"DEBUG: Módulos ativos encontrados:")
    for mod in modulos_com_dados:
        print(f"  Módulo {mod['modulo']} ({mod['posicao']}): {', '.join(mod['cores'])}")
    
    return modulos_ativos

def extrair_info_excel(caminho_excel, caminho_saida_txt, caminho_saida_json):
    """
    Abre um arquivo Excel, encontra uma linha de descrição, analisa-a e
    salva a informação em arquivos de texto e JSON.

    Args:
        caminho_excel (str): O caminho para o arquivo Excel de entrada.
        caminho_saida_txt (str): O caminho para o arquivo de texto de saída.
    """
    # --- Padrões de Regex para diferentes tipos de produtos ---
    # Esta lista pode ser expandida para cobrir mais variações.
    # O padrão mais específico deve vir primeiro.
    regex_patterns = [
        # Padrão para PRIMUS (FORCE, SALIENT, PRIME, ULTRA)
        re.compile(
            r"(?P<codigo_firmware>[\d\w\.]+)\s*-\s*SIN VISUAL\s+"
            r"(?P<produto>PRIMUS|ARES)\s+(?P<modelo>FORCE|SALIENT|PRIME|ULTRA)\s+"
            r"(?P<tamanho>\d+P)\s+(?P<modulos>\d+M)\s+(?P<leds>\d+L)\s+"
            r"(?P<cores_leds>\(.*?\))\s+(?P<funcoes_extras>\w+)\s+"
            r"TP\s+(?P<cores_tampas>[\w/]+)",
            re.IGNORECASE
        ),
        # Padrão para Colimador (ECO, STAR)
        re.compile(
            r"(?P<codigo_firmware>[\d\w\.]+)\s*-\s*SIN VISUAL\s+"
            r"(?P<produto>PRIMUS|ARES)\s+(?P<modelo>ECO|STAR)\s+"
            r"(?P<tamanho>\d+P)\s+(?P<modulos>\d+M)\s+(?P<leds>\d+L)\s+"
            r"(?P<cores_leds>\(.*?\))\s+(?P<funcoes_extras>\w+)\s+"
            r"TP\s+(?P<cores_tampas>[\w/]+)",
            re.IGNORECASE
        ),
        # Padrão genérico (fallback)
        re.compile(r"(?P<codigo_firmware>[\d\w\.]+)\s*-\s*SIN VISUAL\s+(?P<descricao_completa>.+)", re.IGNORECASE)
    ]

    # Verifica se o arquivo Excel existe antes de continuar
    if not os.path.exists(caminho_excel):
        print(f"Erro: O arquivo '{caminho_excel}' não foi encontrado.")
        return

    try:
        # Carrega o workbook (o arquivo Excel como um todo)
        workbook = openpyxl.load_workbook(caminho_excel, read_only=True)

        # Prioriza a planilha "47 POL", depois busca por qualquer "POL"
        sheet_name = None
        sheet_names_upper = [s.upper() for s in workbook.sheetnames]
        
        if "47 POL" in sheet_names_upper:
            sheet_name = workbook.sheetnames[sheet_names_upper.index("47 POL")]
        else:
            for name in workbook.sheetnames:
                if "POL" in name.upper():
                    sheet_name = name
                    print(f"AVISO: Planilha '47 POL' não encontrada. Usando a primeira alternativa: '{sheet_name}'")
                    break
        
        if not sheet_name:
            sheet_name = workbook.sheetnames[0] # Usa a primeira se nenhuma "POL" for encontrada
            print(f"AVISO: Nenhuma planilha com 'POL' encontrada. Usando a primeira: '{sheet_name}'")
        
        sheet = workbook[sheet_name]    

        # Procura pela célula que contém a string de descrição
        info_encontrada = None
        celula_encontrada = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "SIN VISUAL" in cell.value.upper():
                    info_encontrada = cell.value
                    celula_encontrada = cell.coordinate
                    break
            if info_encontrada:
                break

        if info_encontrada:
            print(f"Informação encontrada na célula {celula_encontrada}: {info_encontrada}")
            
            # Analisa a string para extrair os dados
            dados_estruturados = parse_info_string(info_encontrada, regex_patterns)

            # Se os dados foram extraídos, adiciona o mapeamento de cores
            if dados_estruturados:
                dados_estruturados["mapeamento_cores_modulos"] = mapear_cores_posicoes_fixas(sheet)
                
                # Recalcula o número real de módulos baseado nos dados encontrados
                modulos_reais = contar_modulos_ativos(dados_estruturados["mapeamento_cores_modulos"])
                dados_estruturados["quantidade_modulos_original"] = dados_estruturados.get("quantidade_modulos", "0")
                dados_estruturados["quantidade_modulos"] = str(modulos_reais)
                dados_estruturados["quantidade_modulos_ativos"] = modulos_reais
                
                print(f"Módulos encontrados na descrição: {dados_estruturados['quantidade_modulos_original']}")
                print(f"Módulos ativos encontrados na planilha: {modulos_reais}")

            # Salva o arquivo de texto (.txt)
            with open(caminho_saida_txt, 'w', encoding='utf-8') as f_txt:
                if dados_estruturados:
                    print(f"Salvando dados estruturados no arquivo '{caminho_saida_txt}'...")
                    f_txt.write("Dados Extraídos do Excel:\n")
                    f_txt.write("="*30 + "\n")
                    for chave, valor in dados_estruturados.items():
                        f_txt.write(f"{chave.replace('_', ' ').title()}: {valor}\n")
                else:
                    f_txt.write(f"String original encontrada: {info_encontrada}\n")
                    f_txt.write("Não foi possível analisar a string para extrair os dados estruturados.")
            
            # Salva o arquivo JSON (.json)
            with open(caminho_saida_json, 'w', encoding='utf-8') as f_json:
                if dados_estruturados:
                    print(f"Salvando JSON no arquivo '{caminho_saida_json}'...")
                    json.dump(dados_estruturados, f_json, indent=4, ensure_ascii=False)
                    print("Arquivos gerados com sucesso!")
                else:
                    json.dump({"erro": "Não foi possível analisar a string", "string_original": info_encontrada}, f_json, indent=4, ensure_ascii=False)
                    print(f"AVISO: Não foi possível analisar a string. Verifique o padrão. Arquivo JSON de erro gerado.")
        else:
            mensagem_erro = f"Nenhuma célula com 'SIN VISUAL' foi encontrada na planilha '{sheet_name}'."
            print(mensagem_erro)
            with open(caminho_saida_txt, 'w', encoding='utf-8') as f:
                f.write(mensagem_erro)
            with open(caminho_saida_json, 'w', encoding='utf-8') as f:
                json.dump({"erro": mensagem_erro}, f, indent=4, ensure_ascii=False)

    except Exception as e:
        print(f"Ocorreu um erro ao processar o arquivo Excel: {e}")

# --- Como usar a função ---

def main():
    """
    Função principal que processa argumentos da linha de comando
    """
    if len(sys.argv) < 2:
        print("Uso: python excel_parser.py <caminho_do_arquivo_excel> [caminho_saida_txt] [caminho_saida_json]")
        print("Exemplo: python excel_parser.py arquivo.xlsm dados.txt dados.json")
        sys.exit(1)
    
    # Pega o caminho do arquivo Excel dos argumentos
    nome_arquivo_excel = sys.argv[1]
    
    # Define os arquivos de saída (usa argumentos se fornecidos, senão usa padrões)
    nome_arquivo_txt = sys.argv[2] if len(sys.argv) > 2 else "dados_extraidos.txt"
    nome_arquivo_json = sys.argv[3] if len(sys.argv) > 3 else "dados_extraidos.json"
    
    # Executa a extração
    extrair_info_excel(nome_arquivo_excel, nome_arquivo_txt, nome_arquivo_json)

# Executa a função principal apenas se o script for executado diretamente
if __name__ == "__main__":
    main()
